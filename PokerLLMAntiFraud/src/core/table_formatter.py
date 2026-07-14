import os
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter
from .mydataclasses import FraudRecord
from typing import List

class TableFormatter:
    def __init__(self, filename: str = None):
        if filename is None:
            project_root = Path(__file__).parent.parent.parent
            filename = str(project_root / "results.xlsx")
        self.filename = filename
        base, ext = os.path.splitext(filename)
        self.temp_filename = f"{base}_temp{ext}"

    def change_file_name(self, new_filename: str) -> None:
        self.filename = new_filename
        base, ext = os.path.splitext(new_filename)
        self.temp_filename = f"{base}_temp{ext}"

    def save_result(self, record: FraudRecord) -> None:
        if record is None:
            return

        # 1. Try to merge any leftover temp file into the main file
        self._merge_temp_if_possible()

        # 2. Try to write new record directly to the main file
        try:
            wb, ws = self._prepare_workbook(self.filename)
            self._insert_records(ws, [record])  # передаём список из одной записи
            self._auto_fit_columns(ws)
            wb.save(self.filename)
            return
        except PermissionError:
            pass

        # 3. Fallback: append new records to the temp file
        wb, ws = self._prepare_workbook(self.temp_filename)
        self._insert_records(ws, [record])
        self._auto_fit_columns(ws)
        wb.save(self.temp_filename)

    def flush(self) -> None:
        """Attempt to move data from temp file to main file (call on shutdown)."""
        self._merge_temp_if_possible()

    def _merge_temp_if_possible(self) -> None:
        """If temp file exists and main file is writable, merge temp data into main."""
        if not os.path.exists(self.temp_filename):
            return
        try:
            # Load temp workbook
            temp_wb = openpyxl.load_workbook(self.temp_filename)
            temp_ws = temp_wb.active
            # Load main workbook (or create new)
            main_wb, main_ws = self._prepare_workbook(self.filename)
            # Copy all data rows from temp (skip header) into main at top
            temp_rows = list(temp_ws.iter_rows(min_row=2, values_only=True))
            if temp_rows:
                # Insert empty rows at the top of main sheet
                main_ws.insert_rows(2, len(temp_rows))
                for i, row_data in enumerate(reversed(temp_rows)):
                    for col_idx, value in enumerate(row_data, start=1):
                        main_ws.cell(row=2 + i, column=col_idx, value=value)
                self._auto_fit_columns(main_ws)
                main_wb.save(self.filename)
            temp_wb.close()
            # Remove temp file after successful merge
            os.remove(self.temp_filename)
        except PermissionError:
            # Main file still locked, keep temp for next attempt
            pass

    def _prepare_workbook(self, path: str):
        """Opens existing workbook or creates a new one with headers."""
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ['time', 'game_id', 'incident_types', 'player_nicknames', 'description']
            ws.append(headers)
        return wb, ws

    def _insert_records(self, ws, records: List[FraudRecord]) -> None:
        """Insert records at the top of the sheet (after header)."""
        num = len(records)
        ws.insert_rows(2, num)
        for i, record in enumerate(reversed(records)):
            row_data = self._record_to_row(record)
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=2 + i, column=col_idx, value=value)

    def _record_to_row(self, record: FraudRecord) -> list:
        time_str = record.time.strftime('%Y-%m-%d %H:%M:%S')
        game_id_str = str(record.game_id)
        types_str = ', '.join(record.incident_types) if record.incident_types else ''
        players_str = ', '.join(map(str, record.participants_ids)) if record.participants_ids else ''
        description_str = record.description
        return [time_str, game_id_str, types_str, players_str, description_str]

    def _auto_fit_columns(self, ws) -> None:
        """Set column width based on the string length."""
        for col_idx in range(1, 6):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                for cell_value in row:
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
            adjusted_width = max_length + 5
            ws.column_dimensions[col_letter].width = adjusted_width